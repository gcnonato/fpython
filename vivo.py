from datetime import datetime

import os
import random
from time import sleep

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import WHITE

from decouple import config
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as CondicaoExperada
from selenium.webdriver.support.ui import WebDriverWait

url = 'https://login.vivo.com.br/loginmarca/appmanager/marca/publico?origem=https://login.vivo.com.br/saml2/idp/sso/login-return#'

class VivoWithSelenium:
    def __init__(self):
        if os.name != "posix":  # Windows
            self.driver = webdriver.Chrome()
            chrome_options = webdriver.ChromeOptions()
            # chrome_options.add_argument('--headless')
        else:
            self.driver = webdriver.Firefox()
            self.driver.set_window_size(1120, 550)
        self.wait = WebDriverWait(
            self.driver,
            10,
            poll_frequency=1,
            ignored_exceptions=[
                NoSuchElementException,
                ElementNotVisibleException,
                ElementNotSelectableException,
            ],
        )

    @staticmethod
    def type_like_a_person(sentence, single_input_field):
        """ Essa função basicamente simulará a digitação de uma pessoa"""
        print("going to start typing message into message share text area")
        for letter in sentence:
            single_input_field.send_keys(letter)
            sleep(random.randint(1, 10) / 30)

    def percorrer_tabela(self, pagina, first_page_previous, _f):
        driver = self.driver
        quantidade_linha_tabela = len(driver.find_elements_by_xpath('//tbody//tr'))
        first_page_current = driver.find_elements_by_xpath('//tbody//tr')[0].text
        print(f'Primeira Página Anterior..: {first_page_previous}\n'
              f'Primeira Página Atual..: {first_page_current}\n'
              f'Página..: {pagina} '
              f'- Quantidade de linhas...: {quantidade_linha_tabela}')
        pagina += 1
        if first_page_current not in first_page_previous:
            for row in driver.find_elements_by_xpath('//tbody//tr'):
                _f.write(row.text)
                print(row.text)
                _f.write('\n')
            print('*' * 60)
            return pagina, quantidade_linha_tabela, first_page_current
        first_page_current = False
        return pagina, quantidade_linha_tabela, first_page_current

    def main(self, url, params):
        fileout = f"vivo3g-{datetime.today().date()}-{datetime.today().time().isoformat().split('.')[0].replace(':','_')}.txt"
        homepath = os.path.expanduser(os.getenv('USERPROFILE'))
        desktoppath = 'Desktop'
        archive = os.path.join(homepath, desktoppath, fileout)
        with open(archive, 'w') as _f:
            driver = self.driver
            driver.get(url)
            xpath_clicar_button_login = '//*[@id="main-menu"]/div[1]/nav/div/a[2]'
            clicar_button_login = self.wait.until(
                CondicaoExperada.element_to_be_clickable(
                    (By.XPATH, xpath_clicar_button_login)
                )
            )
            clicar_button_login.click()
            sleep(random.randint(5, 7) / 30)
            xpath_vivo_login_username = '//input[@name="vivo_login_username"]'
            input_email = self.wait.until(
                CondicaoExperada.element_to_be_clickable(
                    (By.XPATH, xpath_vivo_login_username)
                )
            )
            input_email.clear()
            self.type_like_a_person(params["cpf"],input_email)
            sleep(random.randint(5, 7) / 30)
            xpath_input_password = '//input[@name="vivo_login_password"]'
            input_password = self.wait.until(
                CondicaoExperada.element_to_be_clickable(
                    (By.XPATH, xpath_input_password)
                )
            )
            sleep(random.randint(5, 7) / 30)
            input_password.clear()
            self.type_like_a_person(params["password"],input_password)
            sleep(random.randint(15, 18) / 30)
            xpath_button_login = \
                "/html/body/div[2]/div[2]/header/div/div/div/div/div[1]/div[2]/div/div/div[2]/form/button"
            button_login = self.wait.until(
                CondicaoExperada.element_to_be_clickable(
                    (By.XPATH, xpath_button_login)
                )
            )
            button_login.click()
            sleep(15)
            close_banner = driver \
                .find_element_by_class_name('fecharAlert')
            try:
                close_banner.click()
            except ElementNotInteractableException as err:
                print('Sem banner para fechar')
            # Até aqui conecta no site
            sleep(5)
            # fim_do_ciclo = driver.find_elements_by_xpath('//strong[@data-bind]')[1].text
            url_extrato_detalhado = 'https://meuvivo.vivo.com.br/content/vivo/meu-vivo/meu-consumo/consumo.html'
            driver.get(url_extrato_detalhado)
            sleep(5)
            consumido = driver.find_element_by_xpath('//span[@data-bind="text:tabledata.Consumption"]').text
            consumido_em_porcentagem = driver.find_element_by_xpath('//span[@data-bind="text:tabledata.ConsumptionPercent"]').text
            disponivel = driver.find_element_by_xpath('//span[@data-bind="text:tabledata.available"]').text
            disponivel_em_porcentagem = driver.find_element_by_xpath('//span[@data-bind="text:tabledata.availablePercent"]').text


            sleep(5)
            link_extrato_mais_detalhado = driver.find_element_by_xpath('//a[@data-content-id="impl_see_detailed_extract_data"]')
            link_extrato_mais_detalhado.click()
            sleep(5)
            pagina = 1
            first_page_previous = ''
            print('Inicio da tabela')
            _f.write(f'Consumido..: {consumido} em %..: {consumido_em_porcentagem}\n')
            _f.write(f'Disponível..: {disponivel} em %..: {disponivel_em_porcentagem}\n')
            _f.write(f'{"*"*80}\n')
            pagina, quantidade_linha_tabela, first_page_previous = \
                self.percorrer_tabela(pagina, first_page_previous, _f)
            sleep(5)
            while True:
                driver.find_element_by_xpath('//a[@title="Próxima página"]').click()
                sleep(2)
                pagina, quantidade_linha_tabela, first_page_previous = \
                    self.percorrer_tabela(pagina, first_page_previous, _f)
                if quantidade_linha_tabela < 5 or not first_page_previous:
                    break
            driver.close()
            print('Fechando o arquivo...')
            print(f'Consumido..: {consumido} em %..: {consumido_em_porcentagem}\n'
                  f'Disponível..: {disponivel} em %..: {disponivel_em_porcentagem}\n')
            print('Fim da tabela')


def load_txt():
    fileout = "gasto_3g.txt"
    homepath = os.path.expanduser(os.getenv('USERPROFILE'))
    desktoppath = 'Desktop'
    archive = os.path.join(homepath, desktoppath, fileout)
    print('Abrindo arquivo...')
    with open(archive, 'r') as _f:
        file_read = _f.readlines()
    wb = Workbook()
    ws = wb.active
    nro_char = 0
    file_name = "formula.xlsx"
    path = os.path.join(homepath, desktoppath, file_name)
    # try:
    #     fi = open(path, "r")
    #     # perform file operations
    # except PermissionError as err:
    #     fi.close()
    #     print(err)
    #     return
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    first_time = True
    qt_mensal = 0
    for info in file_read:
        # print(info)
        cell_A = f"{''.join(('A', str(nro_char + 1)))}"
        cell_B = f"{''.join(('B', str(nro_char + 1)))}"
        cell_C = f"{''.join(('C', str(nro_char + 1)))}"
        ws[cell_A].font = Font(
            name='Arial',
            size=14,
            color="00000000"
        )
        ws[cell_A].fill = PatternFill(
            fgColor = WHITE,
            fill_type = "solid"
        )
        ws[cell_B].font = Font(
            name='Arial',
            size=14,
            color="00000000"
        )
        ws[cell_B].fill = PatternFill(
            fgColor = WHITE,
            fill_type = "solid"
        )
        if '2020' in info:
            ws[cell_A] = info.strip()
            # Testa se o dia do mês é 6 e se for o 6 e a primeira vez, entra
            if datetime.strptime(info.split(',')[0], "%d/%m/%Y").day == 6\
            and first_time:
                    first_time = False
                    qt_mensal = 3000
            nro_char += 1
        elif 'tulo' in info:
            ws[cell_A] = info.strip().split('Título..: ')[1]
            nro_char += 1
        elif '->' in info:
            print(info.split('->')[0], info.split('->')[1])
            ws[cell_A] = info.split('->')[0] # é colocado Gastei..: 7% OU Restante..: 93%
            if 'MB' in info:
                valor = info.split('->')[1] # 215.83 MB
                valor = valor.split(' MB')[0].strip() # 215.83
                ws[cell_B] = valor.replace(".",",") # 215,83
                cell_sum = f"{''.join(('C', str(nro_char + 1)))}"  # C3
                ws[cell_sum] = qt_mensal - float(valor)
            else:
                valor = info.split('->')[1] # 3 GB
                valor = valor.split(' GB')[0].strip() # 3.0 OU 2.5
                ws[cell_C] = valor.replace(".",",") # 215,83
            nro_char += 1
        elif '-----' in info:
            nro_char += 1
        # elif 'Renova em' in info:
        wb.save(path)


if __name__ == "__main__":
    url = "https://www.vivo.com.br/para-voce"
    params = {}
    params["cpf"] = config("VIVO_CPF")
    params["password"] = config("VIVO_PASSWORD")
    bd = VivoWithSelenium()
    bd.main(url, params)
    # load_txt()
