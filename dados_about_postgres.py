import os
import random
from pprint import pprint

import openpyxl
import psycopg2
from decouple import config
from faker import Faker
# first, import a similar Provider or use the default one
from faker.providers import BaseProvider


# Não usado mas deixado para uso futuro
class Faike:
    def __init__(self):
        self.fake = Faker()

    def load(self):
        # create new provider class
        class MyProvider(BaseProvider):
            def foo(self):
                return "bar"

        # then add new provider to faker instance
        self.fake.add_provider(MyProvider)

        # now you can use:
        print(self.fake.foo())
        # 'bar'

        # fake = Faker(['it_IT', 'pt_BR', 'ja_JP'])
        # for _ in range(10):
        # print(fake.name())


class Connection:
    _db = None

    def __init__(self):
        self._db = psycopg2.connect(
            database=config("DATABASE_DOCKER_POSTGRES"),
            host=config("HOST_DOCKER_POSTGRES"),
            user=config("USER_DOCKER_POSTGRES"),
            password=config("PASSWORD_DOCKER_POSTGRES"),
            port=5436,
        )

    def getDb(self):
        return self._db

    def manipular(self, sql):
        try:
            cur = self._db.cursor()
            cur.execute(sql)
            rs = cur.fetchall()
        except Exception as e:
            return None
        return rs

    def inserir(self, dados):
        title, content, published, tenant_id = dados
        sql = f"INSERT INTO public.about_about(title, content, published, tenant_id)"
        sql += f" VALUES('{title}', '{content}', '{published}', {tenant_id})"
        try:
            cursor = self._db.cursor()
            cursor.execute(sql)
            self._db.commit()
            cursor.close()
        except psycopg2.OperationalError as e:
            raise (f"ERROR..: {e}")

    def consultar(self, sql):
        rs = None
        try:
            cur = self._db.cursor()
            cur.execute(sql)
            rs = cur.fetchall()
        except:
            return None
        return rs

    def proximaPK(self, tabela, chave):
        sql = "select max(" + chave + ") from " + tabela
        rs = self.consultar(sql)
        pk = rs[0][0]
        return pk + 1

    def fechar(self):
        self._db.close()

    def importar(self):
        # os.chdir('andre')
        path = os.getcwd()

        files = []
        # r=root, d=directories, f = files
        print(path)
        for r, d, f in os.walk(path):
            if not d:
                # print(f'r.:{r} - d.:{d} - f.:{f}\n')
                for file in f:
                    if "banco".upper() in file:
                        files.append(os.path.join(r, file))
                        # print(f'directory.: {d}  --  file.:{file}')
        for f in files:
            print(f)
            filename = f
        wb = openpyxl.load_workbook(filename)
        worksheet = wb.active
        print(worksheet)
        excel_data = list()
        name_columns = []
        data = {}
        for value in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            name_columns.append(value)
        for datas in worksheet.iter_rows(
            min_row=2,
            # min_col=4,
            max_col=14,
            max_row=5,
            values_only=True,
        ):
            for value, chave in zip(datas, range(0, 14)):
                key = name_columns[0][chave]
                if "empresa".upper() in key:
                    print("*" * 66)
                    print(f"{key} {value}")
                else:
                    print(f"{key} {value}")

    def selectDB(self, dbSelect):
        # LOCAL
        local = {
            "database": config("DATABASE"),
            "host": config("HOST_LOCAL"),
            "user": config("USER_LOCAL"),
            "password": config("PASSWORD_LOCAL"),
        }

        # return local if dbSelect == "1" else aws

    def listar(con, sql):
        for row in con.consultar(sql):
            pprint(row)

    def writeCSV(df):
        df.to_csv("../../gastos.csv")


if __name__ == "__main__":
    con = Connection()
    dados = []
    faker = Faker("pt-br")
    fakers = faker.profile()
    for i in range(10):
        for key, values in fakers.items():
            if key in ["job", "company"]:
                dados.append(values)
        dados.append(random.randint(0, 1))
        dados.append(random.randint(1, 2))
        print(con.inserir(dados))
        pprint(dados)
        dados = []
        fakers = faker.profile()
    # (1, 'Estou infectado?', '<p>Sintomas: corpo mole</p>', False, 1)
    # title, content, published, tenant_id
    # sql = "SELECT * FROM about_about"
    # for about in con.consultar(sql):
    #     print(about)
